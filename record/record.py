
#import openpyxl
from openpyxl import load_workbook as loadwb
from time import sleep
from os import system


file_dir = r'C:\\Users\\bsnawoke\\Desktop\\a1.xlsx'
#wb = openpyxl.load_workbook(file_dir)
wb = loadwb(file_dir) 
ws = wb[wb.sheetnames[0]]
a = len( tuple(ws.columns)[1] )
#for x in tuple(ws.columns)[1]:
#    print (x.value)
#ws[f'A{a+1}'].value = a
#wb.save(file_dir)
prompt = (f'''q
# a    b    c    d    e
#序号 物资 石油1 维修 白箱
#序号 物资 石油2 弹药 蓝箱
#序号 物资 石油3 NUll 紫箱          
''')
          
'''
#
# param_1: num : 钱 B col 
# param_2: char d1 : 弹药 D col D2
# param_3: char d2 : 维修 D col D3
# param_4: char c1 : 可乐 C col C2
# param_5: char c2 : 妹汁 C col C3
# param_6: char c3 : 炸鸡 C col C4
# param_7: char e : 箱子 E col 白e1 E2 蓝e2 E3 紫e3 E4
# param_quit: char q : quit startswith
# param_help: char help : h startswith
#
'''
print(prompt)

value = input(format("输入:"))
#ws.cell(row=2, column=2, value=value) #第二行第二列 B2
#ws[f'A{a}'].value = value #'a = 5 A5'

while value.startswith("q") != 1:
    # 辅助
    if value.startswith('d'):
        if value == 'd1':
            ws['d2'].value = int(ws['d2'].value) + 1
            wb.save(file_dir)
        elif value == 'd2':
            ws['d3'].value = int(ws['d3'].value) + 1
            wb.save(file_dir)
        else:
            print('wrong input!')
    # 石油
    elif value.startswith('c'): 
        if value == 'c1':
            ws['c2'].value = int(ws['c2'].value) + 1
            wb.save(file_dir)
        elif value == 'c2':
            ws['c3'].value = int(ws['c3'].value) + 1
            wb.save(file_dir)
        elif value == 'c3':
            ws['c4'].value = int(ws['c4'].value) + 1
            wb.save(file_dir)
        else:
            print('wrong input!')
    # 箱子
    elif value.startswith('e'):
        if value == 'e1':
            ws['e2'].value = int(ws['e2'].value) + 1
            wb.save(file_dir)
        elif value == 'e2':
            ws['e3'].value = int(ws['e3'].value) + 1
            wb.save(file_dir)
        elif value == 'e3':
            ws['e4'].value = int(ws['e4'].value) + 1
            wb.save(file_dir)
        else: 
            print('wrong input!')
    # 物资        
    elif value.isdigit():

    # 说明文字    
    elif value.startswith('h'):
        system('cls')
        print(prompt)
    else:
        print('wrong input!')
    value = input(format("下一个输入:"))
print('\r',format('goodbye',('—^20s')) )
for i in range(5):   
    print('\r',format(f'{5-i}秒后关闭','-^17s'),end='',flush=True)
    sleep(1)

        