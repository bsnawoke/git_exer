# -*- coding: utf-8 -*-
"""
Created on Wed Jul 22 00:27:27 2020
@author: bsnawoke
"""
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtGui import QIcon # 图标
from gui import Ui_Form
from openpyxl import load_workbook as loadwb
'''
#
# param_1: num : 钱 B col 
# param_2: char d1 : 弹药 D col D2
# param_3: char d2 : 维修 D col D3
# param_4: char c1 : 可乐 C col C2
# param_5: char c2 : 妹汁 C col C3
# param_6: char c3 : 炸鸡 C col C4
# param_7: char e : 箱子 E col 白e1 E2 蓝e2 E3 紫e3 E4
# param_quit: char q : quit startswith
# param_help: char help : h startswith
#
'''
class MainForm(QMainWindow, Ui_Form):
    def __init__(self,parent=None):
        super(MainForm,self).__init__(parent)
        self.setupUi(self)
        self.init_UI()
        self.Oil_1.clicked.connect(self.oil_add_1)    
        self.Oil_2.clicked.connect(self.oil_add_2)
        self.Oil_3.clicked.connect(self.oil_add_3)      
        self.Repair.clicked.connect(self.repair)   
        self.Ammo.clicked.connect(self.ammo)   
        self.Case_1.clicked.connect(self.case_1)   
        self.Case_2.clicked.connect(self.case_2)
        self.Case_3.clicked.connect(self.case_3)
        self.Add_money.clicked.connect(self.money_money_money)
        self.Money_money_money.returnPressed.connect(self.money_money_money)
        self.Cancel.clicked.connect(self.CancelButton)
        self.signal = ""
        self.a = len( tuple(ws.columns)[1] )
        
    def init_UI(self):
        #self.setGeometry(None,None , 300, 220) # 屏幕位置,尺寸
        self.resize(300,150)
        self.setWindowTitle('给老子算')
        self.setWindowIcon(QIcon('icon.jpg'))   
        self.show()
        
    def oil_add_1(self):
        ws['c2'].value = int(ws['c2'].value) + 1
        wb.save(file_dir)
        self.signal = "c2"
        self.label.setText(format(f"可乐:{ws['c2'].value}",
                                  ' ^10s'))
        
    def oil_add_2(self):  
        ws['c3'].value = int(ws['c3'].value) + 1
        wb.save(file_dir)
        self.signal = "c3"
        self.label.setText(format(f"妹汁:{ws['c3'].value}",
                                  ' ^10s'))
    def oil_add_3(self):  
        ws['c4'].value = int(ws['c4'].value) + 1
        wb.save(file_dir)
        self.signal = "c4"
        self.label.setText(format(f"炸虾:{ws['c4'].value}",
                                  ' ^10s'))        
        
    def repair(self):
        ws['d2'].value = int(ws['d2'].value) + 1
        wb.save(file_dir)
        self.signal = "d2"
        self.label.setText(format(f"维修:{ws['d2'].value}",
                                  ' ^10s'))        
    def ammo(self):
        ws['d3'].value = int(ws['d3'].value) + 1
        wb.save(file_dir)
        self.signal = "d3"
        self.label.setText(format(f"弹药:{ws['d3'].value}",
                                  ' ^10s'))          
    def case_1(self):
        ws['e2'].value = int(ws['e2'].value) + 1
        wb.save(file_dir)
        self.signal = "e2"
        self.label.setText(format(f"白箱:{ws['e2'].value}",
                                  ' ^10s'))          
    def case_2(self):
        ws['e3'].value = int(ws['e3'].value) + 1
        wb.save(file_dir)
        self.signal = "e3"
        self.label.setText(format(f"蓝箱:{ws['e3'].value}",
                                  ' ^10s'))     
    def case_3(self):
        ws['e4'].value = int(ws['e4'].value) + 1
        wb.save(file_dir)
        self.signal = "e4"
        self.label.setText(format(f"紫箱:{ws['e4'].value}",
                                  ' ^10s'))             
    def money_money_money(self):
        num = self.Money_money_money.text()
        if num.isdigit():
            
            ws[f'A{self.a+1}'].value = self.a 
            ws[f'B{self.a+1}'].value = int(num)
            ws["d4"].value = int(self.a)
            self.label.setText \
            (format(f"物资:{ws[f'B{self.a+1}'].value}",' ^10s')) 
            self.Money_money_money.setText("")
            wb.save(file_dir)
            self.signal = "my"
            self.a += 1 
        else:
            self.label.setText(format("输尼玛呢",' ^10s'))
            self.Money_money_money.setText("")

# 撤销按钮
    def CancelButton(self):
        if self.signal == "c2":
            ws['c2'].value = int(ws['c2'].value) - 1
            wb.save(file_dir)
            return self.label.setText(format("可乐-1",' ^10s'))     
        elif self.signal == "c3":
            ws['c3'].value = int(ws['c3'].value) - 1
            wb.save(file_dir)
            return self.label.setText(format("妹汁-1",' ^10s'))
        elif self.signal == "c4":
            ws['c4'].value = int(ws['c4'].value) - 1
            wb.save(file_dir)
            return self.label.setText(format("炸虾-1",' ^10s'))
        elif self.signal == "d2":
            ws['d2'].value = int(ws['d2'].value) - 1
            wb.save(file_dir)
            return self.label.setText(format("维修-1",' ^10s'))
        elif self.signal == "d3":
            ws['d3'].value = int(ws['d3'].value) - 1
            wb.save(file_dir)
            return self.label.setText(format("弹药-1",' ^10s'))
        elif self.signal == "e2":
            ws['e2'].value = int(ws['e2'].value) - 1
            wb.save(file_dir)
            return self.label.setText(format("白箱-1",' ^10s'))
        elif self.signal == "e3":
            ws['e3'].value = int(ws['e3'].value) - 1
            wb.save(file_dir)
            return self.label.setText(format("蓝箱-1",' ^10s'))
        elif self.signal == "e4":
            ws['e4'].value = int(ws['e4'].value) - 1
            wb.save(file_dir)
            return self.label.setText(format("紫箱-1",' ^10s'))
        elif self.signal == "my":
            ws[f'A{self.a}'].value = None
            ws[f'B{self.a}'].value = None
            ws["d4"].value = int(self.a) - 1
            wb.save(file_dir)
            self.a -= 1
            return self.label.setText(format("物资撤销",' ^10s'))
        

### 主函数 ###
if __name__ == '__main__':
    file_dir = r'C:\\Users\\bsnawoke\\Desktop\\a1.xlsx'
    wb = loadwb(file_dir) 
    ws = wb[wb.sheetnames[0]]
    
    app = QApplication(sys.argv) # 初始化设置
    gui_show = MainForm() # 读取类
    gui_show.show() # 显示
    sys.exit(app.exec_()) # 循环
    wb.close()